from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.utils
import pandas as pd

class Excel_generator():
    def __init__(self, path_excel: str):
        self.path_excel = path_excel
        self.false_check = {}

    def remove_default_sheet(self):
        try:
            print(self.path_excel)
            workbook = load_workbook(self.path_excel)
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                workbook.save(self.path_excel)
        except FileNotFoundError:
            pass  # No need to remove if the file doesn't exist yet

    def create_or_load_workbook(self):
        self.remove_default_sheet()
        try:
            workbook = load_workbook(self.path_excel)
        except FileNotFoundError:
            workbook = Workbook()
            workbook.save(self.path_excel)
        return workbook

    def run(self, report: dict, name: str):
        self.false_check[name] = {}
        # Load existing or create a new workbook
        workbook = self.create_or_load_workbook()
        # Check if the sheet already exists with the same name
        if name in workbook.sheetnames:
            # Delete the existing sheet with the same name
            workbook.remove(workbook[name])

        # Create a new sheet and add the DataFrame
        sheet = workbook.create_sheet(title=name)
        df = pd.DataFrame.from_dict(report)
        for r in dataframe_to_rows(df, index=True, header=True):
            sheet.append(r)
        
        # Add a new column at the beginning
        sheet.insert_cols(1)

        # Merge cells for 'Root check'
        sheet.merge_cells('A3:A4')
        sheet['A3'] = 'Root check'
        sheet['D3'].value = sheet['C3'].value
        sheet['D4'].value = sheet['C4'].value
        sheet.delete_cols(3)
        sheet.merge_cells('C3:G3')
        sheet.merge_cells('C4:G4')
        sheet.column_dimensions['A'].width = 15

        # Merge cells for 'Sensor check'
        sheet.merge_cells('A5:A14')
        sheet['A5'] = 'Sensor check'

        # Apply formatting to the newly added sheet
        cells_widths = 30
        cells_heights = 30
        for col in range(2, len(df.columns) + 3):
            column_letter = openpyxl.utils.get_column_letter(col)
            sheet.column_dimensions[column_letter].width = cells_widths

        border_style = Side(style='medium', color='000000')
        border_style2 = Side(style='thin', color='000000')

        for row in sheet.iter_rows():
            for cell in row:
                self.false_check[name][row[1].value] = None if row[1].value is not None else None
                if cell.value == False:
                    cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type='solid')
                    row[1].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
                    cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                    sheet.row_dimensions[cell.row].height = cells_heights
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.value == True:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type='solid')
                    cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                    sheet.row_dimensions[cell.row].height = cells_heights
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.value is not None and cell.value != "":
                    cell.fill = PatternFill(start_color="FFE87C", end_color="FFFF00", fill_type='solid')
                    cell.border = Border(left=border_style2, right=border_style2, top=border_style2, bottom=border_style2)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    sheet.row_dimensions[cell.row].height = cells_heights

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == False:
                    self.false_check[name][row[1].value] = False
                # Convert boolean values to strings
                elif isinstance(cell.value, bool):
                    cell.value = 'True' if cell.value else 'False'

        sheet['B1'] = 'Name of the Sensors  -> '
        sheet['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
        sheet['B1'].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

        sheet['A2'] = 'Level of Checks'
        sheet['A2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
        sheet['A2'].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

        # Save changes to the Excel file
        workbook.save(self.path_excel)
        
    def Create_Summary_Sheet(self):
        workbook = load_workbook(self.path_excel)
        if "Summary" in workbook.sheetnames:
            # Delete the existing sheet with the same name
            workbook.remove(workbook["Summary"])
        summary_sheet = workbook.create_sheet(title="Summary", index=0)
        df = pd.DataFrame.from_dict(self.false_check)
        for r in dataframe_to_rows(df, index=True, header=True):
            summary_sheet.append(r)

        border_style = Side(style='medium', color='000000')
        summary_sheet.delete_rows(2, amount=2)

        for row in summary_sheet.iter_rows():
            for cell in row:
                if cell.value == False:
                    cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type='solid')
                elif cell.value == None:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type='solid')
                cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                summary_sheet.row_dimensions[cell.row].height = 25
                summary_sheet.column_dimensions[cell.column_letter].width = 10
                cell.alignment = Alignment(horizontal='center', vertical='center')  
        summary_sheet.column_dimensions['A'].width = 30 

        workbook.save(self.path_excel)

# Usage Example:
# Create an instance of Excel_generator and use the run method
# eg:
# excel_gen = Excel_generator("path_to_your_excel_file.xlsx")
# report_data = {"column1": [True, False], "column2": [False, True]}
# excel_gen
