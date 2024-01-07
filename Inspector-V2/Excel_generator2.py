
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill




class Excel_generator():
    def __init__(self):
        self.final_excel = {}
        self.titles = []
        self.df = []

    def run(self):
        number_of_columns = len(self.final_excel) if self.final_excel else 0
        # Creating a DataFrame using pandas
        df = pd.DataFrame.from_dict(self.final_excel)
        column_widths = 20

        # Write DataFrame to Excel file with specified column widths
        with pd.ExcelWriter('Excel_Report.xlsx', engine='xlsxwriter') as writer:
             df.to_excel(writer, sheet_name='Sheet1', index=True)
             workbook = writer.book
             worksheet = writer.sheets['Sheet1']

             
             color1_format = workbook.add_format({'bg_color': '#FFCC99'})  # Change color code as needed
             color2_format = workbook.add_format({'bg_color': '#FFFFCC'})  # Change color code as needed

             row_index = 0  # Change this to the desired row index

             '''for col in range(number_of_columns):  # Replace 'number_of_columns' with the actual number of columns in your worksheet
                worksheet.write(row_index + 1, col, '  ')'''


             start_col = 1
             while start_col < len(df.columns):
                end_col = start_col + 7 if start_col + 7 <= len(df.columns) else len(df.columns)

                for col in range(start_col, end_col):
                        worksheet.set_column(col, col, column_widths)  # Set width of columns to 15 (modify as needed)
        
                        if (start_col // 7) % 2 == 0:
                             worksheet.set_column(col, col, column_widths, color1_format)
                        else:
                             worksheet.set_column(col, col, column_widths, color2_format)
                        #worksheet.write(0, col, 'header')
                
            
                # Categorize columns under a heading
                category = 'Category Name'
                first_column = 0  # Index of the first column under the category
                last_column = 2   # Index of the last column under the category

                # Merge cells to create a heading for the category
                #worksheet.merge_range(0, start_col, 0, end_col-1, category)
                '''for row in worksheet.iter_rows(values_only=True):
                    for cell_value in row:
                        print(cell_value)'''
                

                start_col += 7
             worksheet.set_column(0, 0, 35)

        workbook = load_workbook('Excel_Report.xlsx')
        sheet = workbook.active

             # Iterate through all cells in the worksheet and print their values
        for row in sheet.iter_rows(min_row=1, max_row=11, min_col=1, max_col=466):
                 for cell in row:
                     if ( cell.value == False ):
                        red_fill = PatternFill(start_color="FF0000", end_color="000000")  # Create a red fill
                        cell.fill = red_fill        
                          
             
             
    def combine(self,dict,name):
        self.titles.append(name)
        new_excel = name + '_' 
        # Combine dictionaries with modified keys
        self.final_excel.update({new_excel + key: value for key, value in dict.items()})
        #self.final_excel[name] = (dict)


