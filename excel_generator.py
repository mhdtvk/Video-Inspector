
"""!
@file

@brief The `excel_generator` package simplifies the process of generating Excel reports for troubleshooting.

@package excel_generator
This class utilizes the `openpyxl` and `pandas` libraries to create, manipulate, and style Excel workbooks.

"""
## @defgroup excel_generator excel_generator.py
#@{

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment,NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    import os
    import openpyxl.utils
    import pandas as pd
except ImportError as e:
    print(f"Error: {e}")
    exit(1)

class ExcelGenerator:
    """!
    @brief Class for preparing data and generating Excel reports.

    Methods:
    - `__init__(excel_path: str)`: Initializes the ExcelGenerator class.
    - `creating_directory()`: Ensures the existence of the directory for the Excel file.
    - `remove_default_sheet()`: Removes the default 'Sheet' from the Excel file.
    - `create_or_load_workbook()`: Creates a new workbook or loads an existing one.
    - `convert_data_to_excel(workbook, report: dict, name: str)`: Converts data to an Excel sheet.
    - `design_excel()`: Enhances Excel file readability through formatting.
    - `run(report: dict, name: str)`: Processes and saves a report for a specific sensor type.
    - `create_summary_sheet()`: Creates a summary sheet aggregating false checks.
    - `rearrange_critical_rows()`: Rearranges critical rows based on predefined criteria.
    """
    
    def __init__(self, excel_path: str):
        """!
        @brief Initializes the ExcelGenerator class.

        @param excel_path (str): The path where the Excel report should be saved.
        """
        self.excel_file_path = f"{excel_path}_Excel_Report.xlsx"
        self.false_check = {}
        self.df = None
        self.creating_directory()

    def creating_directory(self):
        """! 
        @brief Ensures the directory for the Excel file exists; creates it if not.
        """
        if not os.path.exists(os.path.dirname(self.excel_file_path) ):
              os.makedirs(os.path.dirname(self.excel_file_path))

    def remove_default_sheet(self):
        """
        @brief Removes the default 'Sheet' from the Excel file.
        """

        try:
            workbook = load_workbook(self.excel_file_path)
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                workbook.save(self.excel_file_path)
        except :
            pass  # No need to remove if the file doesn't exist yet

    def create_or_load_workbook(self):
        """
        @brief Creates a new workbook or loads an existing one.
        """


        self.remove_default_sheet()
        try:
            workbook = load_workbook(self.excel_file_path)
        except :
            workbook = Workbook()
            workbook.save(self.excel_file_path)
        return workbook

    def convert_data_to_excel(self,workbook, report: dict, name: str):
        """
        @brief  Converts the provided data dictionary to an Excel sheet in the workbook.
        """


         # Create a new sheet and add the DataFrame
        sheet = workbook.create_sheet(title=name)
        self.df = pd.DataFrame.from_dict(report)

        for r in dataframe_to_rows(self.df, index=True, header=True):
            sheet.append(r)

        return workbook
    
    def design_excel(self) :
        """
        @brief Formats the Excel sheet with color-coding and styling.
        """

        workbook = load_workbook(self.excel_file_path)

        # Select the active sheet
        for sheet_name in workbook.sheetnames:
            if sheet_name != "Summary":
                sheet = workbook[sheet_name]
                # Add a new column at the beginning
                sheet.insert_cols(1)

                sheet.merge_cells('A3:A4')
                sheet['A3'] = 'Root check'
                sheet['D3'].value = sheet['C3'].value
                sheet['D4'].value = sheet['C4'].value
                sheet.delete_cols(3)
                sheet.merge_cells('C3:G3')
                sheet.merge_cells('C4:G4')
                sheet.column_dimensions['A'].width = 18

                sheet.merge_cells('A5:A11')
                sheet['A5'] = 'Sensors check (L1)'

                sheet.merge_cells('A18:A22')
                sheet['A18'] = 'Sensors check (L2)'                

                cells_widths = 20
                cells_heights = 30
                for col in range(2, len(self.df.columns) + 3):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    sheet.column_dimensions[column_letter].width = cells_widths

                border_style = Side(style='medium', color='000000')
                border_style2 = Side(style='thin', color='000000')

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value == False:
                            cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type='solid')
                            row[1].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
                            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                            sheet.row_dimensions[cell.row].height = cells_heights
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell.value == True:
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type='solid')
                            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                            sheet.row_dimensions[cell.row].height = cells_heights
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell.value is not None and cell.value != "":
                            cell.fill = PatternFill(start_color="FFE87C", end_color="FFFF00", fill_type='solid')
                            cell.border = Border(left=border_style2, right=border_style2, top=border_style2, bottom=border_style2)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            sheet.row_dimensions[cell.row].height = cells_heights

                sheet['B1'] = 'Name of the Sensors  -> '
                sheet['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
                sheet['B1'].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

                sheet['A2'] = 'Level of Checks'
                sheet['A2'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
                sheet['A2'].border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                sheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

                sheet.column_dimensions['B'].width = 50

        workbook.save(self.excel_file_path)

    def run(self, report: dict, name: str):
        """
        @brief Processes and saves the report for a specific sensor type.

        @param report (dict): The data dictionary to be processed.
        @param name (str): The name of the sensor type.
        """
        self.false_check[name] = {}
        # Load existing or create a new workbook
        workbook = self.create_or_load_workbook()
        # Check if the sheet already exists with the same name
        if name in workbook.sheetnames:
            # Delete the existing sheet with the same name
            workbook.remove(workbook[name])

        workbook = self.convert_data_to_excel(workbook, report, name)
        sheet = workbook[name]

        for row in sheet.iter_rows():
            self.false_check[name][row[0].value] = True
            for cell in row:
                if cell.value is False:
                    self.false_check[name][row[0].value] = False 
        
        # Save changes to the Excel file
        workbook.save(self.excel_file_path)

    def create_summary_sheet(self):
        """
        @brief Creates a summary sheet aggregating false checks.
        """

        workbook = load_workbook(self.excel_file_path)
        if "Summary" in workbook.sheetnames:
            # Delete the existing sheet with the same name
            workbook.remove(workbook["Summary"])
        summary_sheet = workbook.create_sheet(title="Summary", index=0)
        df = pd.DataFrame.from_dict(self.false_check)
        
        for r in dataframe_to_rows(df, index=True, header=True):
            summary_sheet.append(r)

        summary_sheet.delete_rows(2, amount=2)
        border_style = Side(style='medium', color='000000')

        for row in summary_sheet.iter_rows():
            for cell in row:
                if cell.value == False:
                    cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type='solid')
                    row[0].fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type='solid')

                elif cell.value is True:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type='solid')

                cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                summary_sheet.row_dimensions[cell.row].height = 30
                summary_sheet.column_dimensions[cell.column_letter].width = 18
                cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet.column_dimensions['A'].width = 50

        workbook.save(self.excel_file_path)

    def rearrange_critical_rows(self):
        """
        @brief Rearranges critical rows based on predefined criteria.
        """

        # Load the Excel file
        workbook = load_workbook(self.excel_file_path)

        # Select the active sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract the critical rows and rearrange them (replace with your logic)
            important_rows = []

            for row in range(1, sheet.max_row + 1):
                if (
                    sheet.cell(row=row, column=1).value == 'NumberOfFramesCheck'

                ):
                    important_rows.append(sheet[row])
                    sheet.delete_rows(row)

                if (
                    sheet.cell(row=row, column=2).value == 'NumberOfFramesCheck'
                ):
                    important_rows.append(sheet[row])
                    sheet.delete_rows(row)

            
                # Perform your logic to identify important rows
                # For example, let's say rows with certain criteria are considered important
            for row in range(1, sheet.max_row + 1):
                if (
                    sheet.cell(row=row, column=1).value == 'flexx2_ir_depth_frame_number_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'flexx2_ir_depth_timestamps_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'kinect_ir_depth_frame_number_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'kinect_ir_depth_timestamps_consistency_check'

                ):
                    important_rows.append(sheet[row])
                    sheet.delete_rows(row)

                if (
                    sheet.cell(row=row, column=1).value == 'flexx2_ir_depth_frame_number_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'flexx2_ir_depth_timestamps_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'kinect_ir_depth_frame_number_consistency_check'
                    or sheet.cell(row=row, column=1).value == 'kinect_ir_depth_timestamps_consistency_check'
                ):
                    important_rows.append(sheet[row])
                    sheet.delete_rows(row)

            # Rearrange the important rows in a specific order (replace with your logic)
            # For example, moving critical rows to the top of the sheet
            sheet.append([])
            for row_data in important_rows:
                sheet.append(row_data)

        # Save the modified workbook
        workbook.save(self.excel_file_path)

# @} ##